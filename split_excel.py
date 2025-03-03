import openpyxl
import xlwt
import time
import os

# 用户输入Excel文件路径和存储路径
excel_add = input('请粘贴待拆分表的路径（包含文件名）：')
save_add = input('请粘贴存取路径：')

# 确保存储目录存在
if not os.path.exists(save_add):
    os.makedirs(save_add)

# 读取Excel文件
xlsx = openpyxl.load_workbook(excel_add, data_only=True)
table = xlsx.active  # 获取第一个工作表

# 获取总列数
total_cols = table.max_column  # 总列数

# ** 计算实际非空的总行数 **
total_rows = 1  # 先包含表头
for row in table.iter_rows(min_row=2, values_only=True):  # 跳过表头
    if any(row):  # 只要该行有非空值
        total_rows += 1

print(f"实际有效行数（包含表头）：{total_rows}")

# 让用户输入每个文件的数据行数（不包含表头）
split_size = int(input('请输入每个拆分文件包含的最大数据行数（不包括表头）：'))

# 计算文件数量（向上取整）
num_files = (total_rows - 1 + split_size - 1) // split_size

print(f"预计拆分出的表格数量：{num_files}")

# 读取所有有效数据
data = [row for row in table.iter_rows(values_only=True) if any(row)]  # 过滤空行

# 拆分逻辑
for file_index in range(num_files):
    start_row = 1 + file_index * split_size  # 数据起始行
    end_row = min(start_row + split_size, total_rows)  # 数据结束行

    if start_row >= total_rows:  # 防止创建空文件
        break

    # 创建新Excel文件
    xlsx2 = xlwt.Workbook()
    sheetq = xlsx2.add_sheet('Sheet1')

    # 复制表头
    for col in range(total_cols):
        sheetq.write(0, col, data[0][col])

    # 复制数据
    for row_offset, row_index in enumerate(range(start_row, end_row)):
        for col in range(total_cols):
            sheetq.write(row_offset + 1, col, data[row_index][col])

    # 生成文件名
    file_name = os.path.join(save_add, f"拆分表_{file_index + 1}_{time.strftime('%Y-%m-%d')}.xls")
    xlsx2.save(file_name)
    print(f'已生成：{file_name}（包含 {end_row - start_row} 行数据）')

print("拆分完成！")